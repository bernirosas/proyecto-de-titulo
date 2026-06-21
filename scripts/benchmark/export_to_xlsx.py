"""Exporta los resultados del benchmark a un Excel con varias hojas.

Hojas:
  - `Por_query`: una fila por (query × técnica × k) con mean_grade, P@k,
    n_judged, etc. Es la fuente de verdad — el resto deriva con fórmulas.
  - `Latencias`: una fila por (query × técnica) con la latencia ms.
  - `Resumen`: agregación cross-query (mean / p50 / p95) por técnica × k.
    Las celdas usan AVERAGEIFS / MEDIAN sobre `Por_query` para que sigan
    funcionando si se edita el detalle.
  - `Comparativo`: pares sparse ↔ rrf_sparse con deltas, vía fórmulas
    contra `Resumen`.

Uso (desde la raíz del repo):
    python scripts/benchmark/export_to_xlsx.py \\
        --evaluations-dir scripts/benchmark/evaluations \\
        --output benchmark_results.xlsx
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HYBRID_PREFIX = "rrf_"
K_VALUES = (1, 3, 5, 10)

HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
BODY_FONT = Font(name="Arial", size=10)
BODY_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
BODY_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
HYBRID_FILL = PatternFill("solid", start_color="EEF5FB")


def is_hybrid(technique: str) -> bool:
    return technique.startswith(HYBRID_PREFIX)


def load_evaluations(evaluations_dir: Path) -> list[dict]:
    out = []
    for f in sorted(evaluations_dir.glob("q*_eval.json")):
        out.append(json.loads(f.read_text(encoding="utf-8")))
    return out


def _style_header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_body_cell(cell, *, numeric: bool, hybrid_row: bool) -> None:
    cell.font = BODY_FONT
    cell.alignment = BODY_ALIGN_RIGHT if numeric else BODY_ALIGN_LEFT
    cell.border = THIN_BORDER
    if hybrid_row:
        cell.fill = HYBRID_FILL


def _autosize(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def write_por_query(ws, evaluations: list[dict]) -> tuple[list[str], int]:
    """Por query × técnica × k. Devuelve (lista_técnicas, ultima_fila_datos)."""
    headers = [
        "query_id",
        "tecnica",
        "familia",
        "k",
        "mean_grade",
        "precision_at_k",
        "n_judged",
        "n_missing",
        "n_failed",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    techs_seen: set[str] = set()
    row = 2
    for ev in evaluations:
        qid = ev["query_id"]
        for tech, by_k in ev.get("metrics", {}).items():
            techs_seen.add(tech)
            for k in K_VALUES:
                top = by_k.get(f"top_{k}", {})
                ws.cell(row=row, column=1, value=qid)
                ws.cell(row=row, column=2, value=tech)
                ws.cell(row=row, column=3, value="hibrido" if is_hybrid(tech) else "sparse")
                ws.cell(row=row, column=4, value=k)
                ws.cell(row=row, column=5, value=top.get("mean_grade"))
                ws.cell(row=row, column=6, value=top.get("precision_at_k"))
                ws.cell(row=row, column=7, value=top.get("n_judged"))
                ws.cell(row=row, column=8, value=top.get("n_missing"))
                ws.cell(row=row, column=9, value=top.get("n_failed"))
                hybrid = is_hybrid(tech)
                for c in range(1, 10):
                    _style_body_cell(
                        ws.cell(row=row, column=c),
                        numeric=(c >= 4),
                        hybrid_row=hybrid,
                    )
                row += 1
    ws.freeze_panes = "C2"
    _autosize(ws)
    return sorted(techs_seen), row - 1


def write_latencias(ws, evaluations: list[dict]) -> int:
    headers = ["query_id", "tecnica", "familia", "latencia_ms"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    row = 2
    for ev in evaluations:
        qid = ev["query_id"]
        for tech, lat in (ev.get("latencies") or {}).items():
            ws.cell(row=row, column=1, value=qid)
            ws.cell(row=row, column=2, value=tech)
            ws.cell(row=row, column=3, value="hibrido" if is_hybrid(tech) else "sparse")
            ws.cell(row=row, column=4, value=lat)
            hybrid = is_hybrid(tech)
            for c in range(1, 5):
                _style_body_cell(
                    ws.cell(row=row, column=c),
                    numeric=(c == 4),
                    hybrid_row=hybrid,
                )
            row += 1
    ws.freeze_panes = "C2"
    _autosize(ws)
    return row - 1


def write_resumen(
    ws,
    techs: list[str],
    por_query_last_row: int,
    lat_last_row: int,
) -> None:
    """Cross-query mean / p50 / p95 por técnica × k usando fórmulas sobre Por_query.

    Las columnas calidad/precisión se computan con AVERAGEIFS / fórmulas array
    en lugar de hardcodear, así si la hoja `Por_query` cambia (re-corres
    el benchmark) el resumen se actualiza solo.
    """
    headers = ["tecnica", "familia"]
    for k in K_VALUES:
        headers += [f"mean_grade@{k} (mean)", f"P@{k} (mean)"]
    headers += ["latencia p50 (ms)", "latencia mean (ms)"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    pq_range_q = f"Por_query!$A$2:$A${por_query_last_row}"
    pq_range_t = f"Por_query!$B$2:$B${por_query_last_row}"
    pq_range_k = f"Por_query!$D$2:$D${por_query_last_row}"
    pq_range_mg = f"Por_query!$E$2:$E${por_query_last_row}"
    pq_range_p = f"Por_query!$F$2:$F${por_query_last_row}"
    lat_range_t = f"Latencias!$B$2:$B${lat_last_row}"
    lat_range_v = f"Latencias!$D$2:$D${lat_last_row}"
    del pq_range_q  # not used; AVERAGEIFS works without query filter

    row = 2
    for tech in techs:
        hybrid = is_hybrid(tech)
        ws.cell(row=row, column=1, value=tech)
        ws.cell(row=row, column=2, value="hibrido" if hybrid else "sparse")
        col = 3
        for k in K_VALUES:
            # mean_grade@k (mean cross-query): AVERAGEIFS(mg, tech_col, tech, k_col, k)
            ws.cell(
                row=row,
                column=col,
                value=(
                    f'=IFERROR(AVERAGEIFS({pq_range_mg}, {pq_range_t}, "{tech}", '
                    f"{pq_range_k}, {k}), NA())"
                ),
            )
            col += 1
            ws.cell(
                row=row,
                column=col,
                value=(
                    f'=IFERROR(AVERAGEIFS({pq_range_p}, {pq_range_t}, "{tech}", '
                    f"{pq_range_k}, {k}), NA())"
                ),
            )
            col += 1
        # latencia p50 / mean — necesita array formula MEDIAN(IF(...))
        # Excel/LibreOffice modernos soportan implicit array context.
        ws.cell(
            row=row,
            column=col,
            value=(
                f'=IFERROR(_xlfn._xlws.PERCENTILE.INC(IF({lat_range_t}="{tech}",{lat_range_v}),0.5), NA())'
            ),
        )
        col += 1
        ws.cell(
            row=row,
            column=col,
            value=(f'=IFERROR(AVERAGEIFS({lat_range_v}, {lat_range_t}, "{tech}"), NA())'),
        )
        # Format the whole row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            _style_body_cell(cell, numeric=(c >= 3), hybrid_row=hybrid)
            if c >= 3 and c < len(headers) - 1:
                cell.number_format = "0.000"
            elif c >= len(headers) - 1:
                cell.number_format = "0.0"
        row += 1

    ws.freeze_panes = "C2"
    _autosize(ws)


def write_comparativo(ws, techs: list[str]) -> None:
    """Head-to-head sparse vs hybrid usando fórmulas contra `Resumen`."""
    sparse_partners = sorted(t for t in techs if not is_hybrid(t))

    headers = ["sparse_partner"]
    for k in K_VALUES:
        headers += [
            f"sparse mg@{k}",
            f"hibrido mg@{k}",
            f"Δ mg@{k}",
            f"sparse P@{k}",
            f"hibrido P@{k}",
            f"Δ P@{k}",
        ]
    headers += [
        "sparse lat p50 (ms)",
        "hibrido lat p50 (ms)",
        "Δ lat p50",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    # Columnas en Resumen:
    #   A=tecnica, B=familia
    #   3,4 = mg@1, p@1; 5,6 = mg@3, p@3; 7,8 = mg@5, p@5; 9,10 = mg@10, p@10
    #   11 = lat p50, 12 = lat mean
    col_mg = {k: 3 + 2 * i for i, k in enumerate(K_VALUES)}  # 3, 5, 7, 9
    col_p = {k: 4 + 2 * i for i, k in enumerate(K_VALUES)}  # 4, 6, 8, 10
    col_lat_p50 = 3 + 2 * len(K_VALUES)  # 11

    row = 2
    for sk in sparse_partners:
        hk = f"{HYBRID_PREFIX}{sk}"
        if hk not in techs:
            continue
        ws.cell(row=row, column=1, value=sk)

        col = 2
        for k in K_VALUES:
            mg_col_letter = get_column_letter(col_mg[k])
            p_col_letter = get_column_letter(col_p[k])
            sparse_mg = f'=VLOOKUP("{sk}", Resumen!$A:$L, {col_mg[k]}, FALSE)'
            hibrido_mg = f'=VLOOKUP("{hk}", Resumen!$A:$L, {col_mg[k]}, FALSE)'
            sparse_p = f'=VLOOKUP("{sk}", Resumen!$A:$L, {col_p[k]}, FALSE)'
            hibrido_p = f'=VLOOKUP("{hk}", Resumen!$A:$L, {col_p[k]}, FALSE)'
            del mg_col_letter, p_col_letter  # only for clarity

            ws.cell(row=row, column=col, value=sparse_mg)
            col += 1
            ws.cell(row=row, column=col, value=hibrido_mg)
            col += 1
            ws.cell(
                row=row,
                column=col,
                value=f"={get_column_letter(col-1)}{row}-{get_column_letter(col-2)}{row}",
            )
            col += 1
            ws.cell(row=row, column=col, value=sparse_p)
            col += 1
            ws.cell(row=row, column=col, value=hibrido_p)
            col += 1
            ws.cell(
                row=row,
                column=col,
                value=f"={get_column_letter(col-1)}{row}-{get_column_letter(col-2)}{row}",
            )
            col += 1

        sparse_lat = f'=VLOOKUP("{sk}", Resumen!$A:$L, {col_lat_p50}, FALSE)'
        hibrido_lat = f'=VLOOKUP("{hk}", Resumen!$A:$L, {col_lat_p50}, FALSE)'
        ws.cell(row=row, column=col, value=sparse_lat)
        col += 1
        ws.cell(row=row, column=col, value=hibrido_lat)
        col += 1
        ws.cell(
            row=row,
            column=col,
            value=f"={get_column_letter(col-1)}{row}-{get_column_letter(col-2)}{row}",
        )

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            _style_body_cell(cell, numeric=(c >= 2), hybrid_row=False)
            if c >= 2 and c < len(headers) - 2:
                cell.number_format = "0.000"
            elif c >= len(headers) - 2:
                cell.number_format = "0.0"
        row += 1

    ws.freeze_panes = "B2"
    _autosize(ws)


def main(evaluations_dir: Path, output: Path) -> None:
    evaluations = load_evaluations(evaluations_dir)
    if not evaluations:
        raise SystemExit(f"[x] no hay qNNN_eval.json en {evaluations_dir}")

    wb = Workbook()
    wb.remove(wb.active)

    ws_pq = wb.create_sheet("Por_query")
    techs, pq_last = write_por_query(ws_pq, evaluations)

    ws_lat = wb.create_sheet("Latencias")
    lat_last = write_latencias(ws_lat, evaluations)

    ws_res = wb.create_sheet("Resumen")
    write_resumen(ws_res, techs, pq_last, lat_last)

    ws_cmp = wb.create_sheet("Comparativo")
    write_comparativo(ws_cmp, techs)

    # Orden de sheets: Resumen primero (lo que abre el usuario)
    wb.move_sheet("Resumen", offset=-3)
    wb.move_sheet("Comparativo", offset=-2)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"[+] {output}")
    print(f"    {len(evaluations)} queries · {len(techs)} técnicas")
    print(f"    Por_query: {pq_last - 1} filas · Latencias: {lat_last - 1} filas")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--evaluations-dir",
        type=Path,
        default=Path("scripts/benchmark/evaluations"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("scripts/benchmark/benchmark_results.xlsx"),
    )
    args = parser.parse_args()
    main(args.evaluations_dir, args.output)
